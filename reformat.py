# !/urs/bin/python
# -*- coding: utf-8 -*-
import sys
import PyQt5
from PyQt5.QtWidgets import QApplication, QWidget
from PyQt5 import uic
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


class App(QWidget):
    def __init__(self):
        self.start_ui()
        self.set()

    def start_ui(self):
        self.ui = uic.loadUi('untitled.ui')
        self.ui.show()

    def set(self):
        self.ui.commandLinkButton.clicked.connect(lambda: self.reformat_go())

    def reformat_start(self):
        for num in range(3500, 8000):
            try:
                self.reformat_zn(f'Загрузочная накладная {num}.xlsx')
                self.reformat_ml(f'Маршрутный лист {num}.xlsx')
            except:
                pass

    def reformat_zn(self, document_number):
        self.document_number = document_number
        self.open_file = load_workbook(self.document_number)
        self.ws_open = self.open_file.active

        self.ws_open['F2'] = self.search_data()
        self.ws_open['F5'] = self.search_city()

        for num in range(2, 9):
            self.ws_open.row_dimensions[num].height = 37

        for num in range(9, self.length_determination_zn() + 1):
            size_zn = ''
            for letter in ['V', 'AF', 'B', 'C', 'F', 'K', 'O', 'X', 'Z', 'AB']:
                self.ws_open[letter + str(num)].alignment = Alignment(horizontal='center', vertical='center',
                                                                      wrap_text=True)
                if letter == 'V':
                    size_zn = str(self.ws_open[letter + str(num)].value).count('1')
                    if size_zn == 1 or size_zn == 2 or size_zn == 0:
                        size_zn = 125
                    elif size_zn == 3 or size_zn == 4:
                        size_zn = 175
                    elif size_zn == 5 or size_zn == 6:
                        size_zn = 220
                    else:
                        size_zn = 420
                if len(str(self.ws_open[letter + str(num)].value)) > 110 and size_zn < len(
                        str(self.ws_open[letter + str(num)].value)):
                    size_zn = len(str(self.ws_open[letter + str(num)].value))

                self.ws_open[letter + str(num)].font = Font(size=11)
                self.ws_open.row_dimensions[num].height = size_zn

        self.ws_open.column_dimensions['Y'].width = 15
        self.ws_open.column_dimensions['AJ'].width = 20
        # self.ws_open.delete_rows(1)
        self.open_file.save(self.document_number)
        self.open_file.close()

    def reformat_ml(self, document_number):
        self.document_number = document_number
        self.open_file = load_workbook(self.document_number)
        self.ws_open = self.open_file.active

        self.ws_open['F2'] = self.ws_open['T9'].value
        for letter in ['A', 'C', 'F', 'K', 'B', 'P', 'T', 'V', 'X', 'AB', 'AI', 'AK', 'AM', 'AO', 'AS', 'AX', 'BC']:
            for num in range(2, 9):
                try:
                    self.ws_open.row_dimensions[num].height = 45
                    self.ws_open[letter + str(num)].font = Font(size=12)
                except:
                    pass
        for num in range(9, self.length_determination_zn() + 2):
            size_zn = ''
            for letter in ['AI', 'AX', 'A', 'C', 'F', 'K', 'B', 'P', 'T', 'V', 'X', 'AB', 'AK', 'AM', 'AO', 'AS', 'BC']:
                self.ws_open[letter + str(num)].alignment = Alignment(horizontal='center', vertical='center',
                                                                      wrap_text=True)
                if letter == 'AI':
                    size_zn = str(self.ws_open[letter + str(num)].value).count('1')
                    if size_zn == 1 or size_zn == 2 or size_zn == 0:
                        size_zn = 240
                    elif size_zn == 3 or size_zn == 4:
                        size_zn = 300
                    elif size_zn == 5 or size_zn == 6:
                        size_zn = 360
                    else:
                        size_zn = 600

                if len(str(self.ws_open[letter + str(num)].value)) > 110 and size_zn < len(
                        str(self.ws_open[letter + str(num)].value)) * 2.4:
                    size_zn = len(str(self.ws_open[letter + str(num)].value)) * 2.4

                self.ws_open[letter + str(num)].font = Font(size=18)
                self.ws_open.row_dimensions[num].height = size_zn

        self.ws_open.column_dimensions['B'].width = 10
        self.ws_open.column_dimensions['O'].width = 10
        self.ws_open.column_dimensions['S'].width = 12
        self.ws_open.column_dimensions['W'].width = 11
        self.ws_open.column_dimensions['AL'].width = 13
        self.ws_open.column_dimensions['BB'].width = 10
        self.ws_open.column_dimensions['U'].width = 8.5
        self.ws_open.column_dimensions['AW'].width = 15
        self.ws_open.column_dimensions['J'].width = 1
        self.ws_open.column_dimensions['I'].width = 4

        self.open_file.save(self.document_number)
        self.open_file.close()

    def length_determination_zn(self):
        if self.ws_open == self.open_file.active:
            for i in range(9, 150):
                if self.ws_open['K' + str(i)].value == None:
                    break
                else:
                    self.num = i
        return self.num

    def search_data(self):
        for column_index in range(10, 170):
            try:
                if 'ЗН' in self.ws_open['C' + str(column_index)].value:
                    data = self.ws_open['C' + str(column_index)].value[-8::]
                    break
            except:
                pass
        return data

    def search_city(self):
        city_lst = ['Ростов-на-Дону', 'Харьков', 'Анапа', 'Краснодар', 'Киев', 'Белгород', 'Ставрополь',
                    'Санкт-Петербург',
                    'Киев', 'Ейск']
        for column_index in range(10, 100):
            if self.ws_open['H' + str(column_index)].value in city_lst:
                return self.ws_open['H' + str(column_index)].value


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    app.exec_()
